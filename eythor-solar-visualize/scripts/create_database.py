import json
import sqlite3
import os

# Paths
excel_path = r"C:\Users\ayush\Downloads\DataSheet_ComparePV_FINAL.xlsx"
output_dir = r"C:\Users\ayush\Desktop\eythor-intern\eythor-solar-visualize\src\data"
db_path = os.path.join(output_dir, "solarPanels.db")
json_path = os.path.join(output_dir, "solarPanels_detailed.json")

# First convert Excel to JSON using openpyxl
import openpyxl
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["Solar Panel Database"]

headers = []
for cell in ws[1]:
    headers.append(cell.value.strip() if cell.value else "")

panels = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None and row[1] is None:
        continue
    entry = {}
    for i, h in enumerate(headers):
        val = row[i] if i < len(row) else None
        if val is not None:
            val_str = str(val).strip()
            if val_str.lower() in ('-', 'none', 'not found', ''):
                val_str = None
            else:
                try:
                    if '.' in val_str:
                        val_str = float(val_str)
                    else:
                        val_str = int(val_str)
                except (ValueError, TypeError):
                    pass  # Keep as string
        else:
            val_str = None
        
        # Map Excel headers to TypeScript interface field names
        field_map = {
            'Manufacturer': 'manufacturer',
            'Model Name': 'model',
            'Rated Power (Wp)': 'ratedPower',
            'Power Density (Wp/m²)': 'powerDensity',
            'Length (mm)': 'length',
            'Width (mm)': 'width',
            'Thickness (mm)': 'thickness',
            'Weight (kg)': 'weight',
            'Temperature Coefficient of Pmax (%/°C)': 'tempCoeff',
            'Annual Degradation (%)': 'annualDegradation',
            'Source URL': 'sourceUrl'
        }
        
        field_name = field_map.get(h, h)
        entry[field_name] = val_str
    
    if entry.get("model"):
        panels.append(entry)

# Save detailed JSON
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(panels, f, indent=2, ensure_ascii=False)
print(f"Saved {len(panels)} panels to JSON")

# Create SQLite database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Create table with all columns
cursor.execute('''
CREATE TABLE IF NOT EXISTS solar_panels (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    manufacturer TEXT,
    model_name TEXT NOT NULL,
    rated_power REAL,
    power_density REAL,
    length_mm REAL,
    width_mm REAL,
    thickness_mm REAL,
    weight_kg REAL,
    temp_coeff TEXT,
    annual_degradation TEXT,
    source_url TEXT
)
''')

# Insert data
insert_sql = '''
INSERT INTO solar_panels (
    manufacturer, model_name, rated_power, power_density, length_mm, width_mm,
    thickness_mm, weight_kg, temp_coeff, annual_degradation, source_url
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
'''

count = 0
for p in panels:
    try:
        cursor.execute(insert_sql, (
            p.get("manufacturer"),
            p.get("model"),
            p.get("ratedPower"),
            p.get("powerDensity"),
            p.get("length"),
            p.get("width"),
            p.get("thickness"),
            p.get("weight"),
            p.get("tempCoeff"),
            p.get("annualDegradation"),
            p.get("sourceUrl")
        ))
        count += 1
        if count % 1000 == 0:
            print(f"Inserted {count} records...")
    except Exception as e:
        print(f"Error inserting {p.get('Model Name')}: {e}")

conn.commit()

# Create indexes for fast searching
cursor.execute('CREATE INDEX IF NOT EXISTS idx_model_name ON solar_panels(model_name)')
cursor.execute('CREATE INDEX IF NOT EXISTS idx_manufacturer ON solar_panels(manufacturer)')
conn.commit()

# Verify
cursor.execute('SELECT COUNT(*) FROM solar_panels')
total = cursor.fetchone()[0]
print(f"Total records in database: {total}")

# Export the database as base64 for inclusion in the web app
import base64
with open(db_path, 'rb') as f:
    db_base64 = base64.b64encode(f.read()).decode('utf-8')

with open(os.path.join(output_dir, "solarPanelsDbBase64.txt"), 'w') as f:
    f.write(db_base64)

print(f"Database size: {os.path.getsize(db_path)} bytes")
print(f"Base64 size: {len(db_base64)} bytes")

conn.close()
print("Done!")