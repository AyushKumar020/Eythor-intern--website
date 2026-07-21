import json
import openpyxl
import sys

# Path to Excel file
excel_path = r"C:\Users\ayush\Downloads\DataSheet_ComparePV_FINAL.xlsx"
output_path = r"C:\Users\ayush\Desktop\eythor-intern\eythor-solar-visualize\src\data\solarPanelDatabase.json"

# Load workbook
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["Solar Panel Database"]

# Parse headers
headers = []
for cell in ws[1]:
    headers.append(cell.value.strip() if cell.value else "")

# Parse rows
panels = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None and row[1] is None:
        continue
    entry = {}
    for i, h in enumerate(headers):
        val = row[i] if i < len(row) else None
        if val is not None:
            val_str = str(val).strip()
            if val_str.lower() in ('-', 'none', 'not found', ''):
                entry[h] = None
            else:
                # Try to convert numeric values
                try:
                    if '.' in val_str:
                        entry[h] = float(val_str)
                    else:
                        entry[h] = int(val_str)
                except (ValueError, TypeError):
                    entry[h] = val_str
        else:
            entry[h] = None
    # Only include if it has a model name
    if entry.get("Model Name"):
        panels.append(entry)

# Write JSON
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(panels, f, indent=2, ensure_ascii=False)

print(f"Exported {len(panels)} solar panels to {output_path}")

# Also create a searchable index by model name keyword
# Create a simplified version for fast lookups
searchable = []
for p in panels:
    searchable.append({
        "model": p.get("Model Name", ""),
        "manufacturer": p.get("Manufacturer", ""),
        "ratedPower": p.get("Rated Power (Wp)"),
        "length": p.get("Length (mm)"),
        "width": p.get("Width (mm)"),
        "thickness": p.get("Thickness (mm)"),
        "weight": p.get("Weight (kg)"),
        "powerDensity": p.get("Power Density (Wp/m²)"),
        "tempCoeff": p.get("Temperature Coefficient of Pmax (%/°C)"),
        "degradation": p.get("Annual Degradation (%)"),
        "sourceUrl": p.get("Source URL")
    })

searchable_path = r"C:\Users\ayush\Desktop\eythor-intern\eythor-solar-visualize\src\data\solarPanels.json"
with open(searchable_path, 'w', encoding='utf-8') as f:
    json.dump(searchable, f, indent=2, ensure_ascii=False)

print(f"Exported simplified {len(searchable)} panels to {searchable_path}")